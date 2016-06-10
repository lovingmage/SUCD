import sys
from collections import defaultdict
from collections import Counter
import os
import openpyxl
import collections
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np

def reverseDict(originalDict):
	newDict = defaultdict(list)
	
	for k, v in originalDict.items():
		newDict[v].append(k)

	return newDict

if __name__ == '__main__':
		
	FILE1 = sys.argv[1]
	FILE2 = sys.argv[2]
	upperbound = int(sys.argv[3])
	LOG_FILE = sys.argv[4]
	original = {}
	kcore = []

	conv = {1:'B',2:'C',3:'D',4:'E',5:'F',6:'G',7:'H',8:'I',9:'J',10:'K'}

	with open(FILE1) as fd1:
		original = eval(fd1.readline())

	originalDict = collections.OrderedDict(sorted(original.items()))	
	number_of_community = len(np.unique(original.values())) 

	kcore = [line.rstrip('\n\r') for line in open(FILE2)]
	lable_list = []
	
	if(not os.path.exists(LOG_FILE+".xlsx")):
		wb = openpyxl.Workbook()

		sheet = wb.active
		sheet.title = 'Sheet1'
		sheet['A1'] = 'Original'
		sheet['B1'] = 'v=10%'
		sheet['C1'] = 'v=20%'
		sheet['D1'] = 'v=30%'
		sheet['E1'] = 'v=40%'
		sheet['F1'] = 'v=50%'
		sheet['G1'] = 'v=60%'
		sheet['H1'] = 'v=70%'
		sheet['I1'] = 'v=80%'
		sheet['J1'] = 'v=90%'
		sheet['K1'] = 'v=100%'
		sheet['A2'] = 'Total'
		
		
		for i in range(number_of_community):
			sheet['A'+str(i+3)] = 'Community' + str(i)
	
	else:
		wb = openpyxl.load_workbook(LOG_FILE+".xlsx")
		sheet = wb.get_sheet_by_name('Sheet1')
	

#total number of nodes in kcore/resilence core

	node_in_core = len(kcore)
	sheet[conv[upperbound]+'2'] = node_in_core 

	for key in kcore:
		lable_list.append(original[key])

	node_count_core = {}
	for i in range(number_of_community):
		node_count_core[i] = lable_list.count(i)	
			

	reverseDict = reverseDict(originalDict)
	
	for key in range(number_of_community):
		sheet[conv[upperbound] + str(key+3)] = (int(node_count_core[key]) * 1.0) / (int(len(reverseDict[key]) * 1.0))


	wb.save(LOG_FILE+".xlsx")
