'''
   This code is designed and written by Chenghong Wang.
   Do not copy, disclose, or distribute without explicit written permission. 

   Author: 			Chenghong Wang <cwang132@syr.edu>
   
   Instruction:			This program is a python package which is designed to compute the community
				partition by using kcore method. The kore method is to firstly reduce the
				original graph by using kcore method; Then perform community detection method
				on the searched kcore subgraph, then recover the subgraph and the community
				structure.
					
	Usage:			This is a package of kcore community detection, to use all function here, by
				using :

					In Linux Command Line using the following command
					
					$python kcore_comm.py [INPUT_FILE_TYPE] [Kcore Number]
	
	***MODIFIED by Rongqian Zhang Nov.18 2015	
				   Chenghong Wang Dec. 20 2015 : Rewrite the graph read method.
				   Chenghong Wang Jan. 5 2016 : Revised the community detection method by put all disconnect node of kcore into exclusive communities.

	***MODIFIED by Hongyu Duan Mar.3 2016: changed ratio part in function "vote_for_node" and write ratio data into log file
			
	***MODIFIED by Hongyu Duan Mar.24 2016: changed log from writting to a txt to a excel file for better information collecting 				
'''

import community
import networkx as nx
from collections import Counter
import os
import time
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

	
#############################################################################################
#	Function : convert_partition_format( ORIGINAL_PARTITION ):			
#											
#											
#	This function will convert Partition Format from the original format 		
#	to the one which we will used in community similarity comparison process.	
#   The input partition format:
#
#	partition_original = {node1: partition 1, node2: partition 2 ...}
#	The keys are the nodes and the key values are the corresponding partition
#
#	The output partition format:
#
#	new_partition = {partition1 : [node1, node2, ...], partition2: [node8, node9,...]...}
#	The keys are the communities, and the key values are the nodes list	
#	which beloings to the key community.
#############################################################################################
	
def convert_partition_format(original_partition):
	communities = list(set(original_partition.values()))
	new_partition = {}
	for community_part in communities:
        	new_partition[community_part] = []
	#print new_partition
	for nodes in original_partition.keys():
        	new_partition[original_partition[nodes]].append(nodes)
	#print new_partition
	return new_partition



##############################################################################################
#       Function : sort_by_neighbor( Kcore Subgraph H, Oroginal Graph G ):		     
#											     
#											
#       This function will sort the recover nodes based on the inside-kcore neighbors,  
#       the function will receive two graph H and G, and then return the asscending     
#		ordered nodes sequence which need to be recovered.			        
##############################################################################################
	
def sort_by_neighbor(H, G):
	global ratio
        sorted_neighbor = []
        sorted_neighbor_1 = []
        sorted_neighbor_2 = []
        sorted_neighbor_3 = []
        sorted_neighbor_4 = []
        sorted_neighbor_5 = []
        H_node = set(H.nodes())
        G_node = set(G.nodes())
        G_H_node = G_node - H_node

        for node in G_H_node :
                neighbors = set(G.neighbors(node))
               
                if len(neighbors)!=0:

                        ratio = (len(neighbors & H_node) * 1.0)/(len(neighbors)*1.0)

			if ratio == 0:
				sorted_neighbor_1.append(node)

			if ratio > 0 and ratio <= 0.25:
				sorted_neighbor_2.append(node)
				
			if ratio > 0.25 and ratio <= 0.5:
				sorted_neighbor_3.append(node)
			
			if ratio > 0.5 and ratio <= 0.75:
				sorted_neighbor_4.append(node)
			
			if ratio > 0.75:
				sorted_neighbor_5.append(node)
						
                else:
                        sorted_neighbor_1.append(node)

        sorted_neighbor = sorted_neighbor_5 + sorted_neighbor_4 + sorted_neighbor_3 + sorted_neighbor_2 + sorted_neighbor_1

	ratio = []
	ratio.append(len(sorted_neighbor_5))
	ratio.append(len(sorted_neighbor_4))
	ratio.append(len(sorted_neighbor_3))
	ratio.append(len(sorted_neighbor_2))
	ratio.append(len(sorted_neighbor_1))
	
        return sorted_neighbor

##############################################################################################
#       Function : vote_for_node( kcore partition kcore_partition, 								
#				  sort order recover nodes sorted_recover_nodes,
#				  Original Graph G ):                   
#                                                                                            
#                                                                                       
#       This function is the main function to recover the nodes into current partition,
#	It will retuen a recovered partition dictionary, and if some nodes can not be 
#	recovered, it will divide them into -1 community                              
##############################################################################################


def vote_for_node(kcore_partition, sorted_recover_nodes, G):
        for node in sorted_recover_nodes:
                #print int(node)
                vote_list = []
                for neighbor in G.neighbors(node):
                        if neighbor not in kcore_partition:
                                continue
                        else:
                                vote_list.append(kcore_partition[neighbor])
                #print vote_list
                vote_dic = Counter(vote_list)
                num = len(vote_list)
                if num == 0:
                        kcore_partition[node] = 0 - int(node)
                else:
                        #Here we want to select the most common nodes
                        kcore_partition[node] = max(vote_dic,key=vote_dic.get) 
        return kcore_partition

#Optional Debugging Code, Could be removed in the next version.
def detect_recover(filename,k):
    #Read Network files as gml file type, create a networkx graph and use the eisted graph file
	#Random graph may have poor performance, erdos renyi graph doesn't have true community structure
	G = nx.read_gml(filename)
	H = nx.k_core (G, int(k))
	#print len(H.nodes())
	#kcore_partition = kcore_partition(H)
	partition = community.best_partition(H)
	#print partition 
	sorted_recover_nodes = sort_by_neighbor(H, G)
	#print sorted_recover_nodes
	vote_for_node(partition, sorted_recover_nodes, G)
	new_partition = vote_for_node(partition, sorted_recover_nodes, G)
	return convert_partition_format(new_partition)

if __name__ == '__main__':
    #Read Network files as gml file type, create a networkx graph and use the eisted graph file
	#Random graph may have poor performance, erdos renyi graph doesn't have true community structure
    
	#Start log file, create log file and start.
	global ratio
	DATA_FILE = sys.argv[1].split("/")
	FILE_LOG_NAME = "LOG_File_"+ (DATA_FILE[-1]) + ".xlsx"

	#check if a log file already created, if not, create one to store infomration we needed for analysis
	if(not os.path.exists(FILE_LOG_NAME)):
		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = 'Sheet1'
		sheet['A1'] = 'k_Value'
		sheet['B1'] = 'StartTime'
		sheet['C1'] = 'ParseFile'
		sheet['D1'] = 'KcoreSearch'
		sheet['E1'] = 'KcorePartition'
		sheet['F1'] = 'KcoreRecover'
		sheet['G1'] = 'Optimization'
		sheet['H1'] = 'EndTime'
		sheet['I1'] = 'TotalTime'
		sheet['J1'] = 'NodesInKcore'
		sheet['K1'] = 'Modularity'
		sheet['L1'] = 'Ratio1'
		sheet['M1'] = 'Ratio2'
		sheet['N1'] = 'Ratio3'
		sheet['O1'] = 'Ratio4'
		sheet['P1'] = 'Ratio5'
		sheet['Q1'] = 'Similarity'
		wb.save(FILE_LOG_NAME)
	
	#Open the log file to write data
	wb = openpyxl.load_workbook(FILE_LOG_NAME)
	sheet = wb.get_sheet_by_name('Sheet1')
	

	Kcore_Value = int(sys.argv[2])
	sheet['A' + str(Kcore_Value + 3)] = Kcore_Value
	
	FILE_PATH = sys.argv[1]
	#start to read data from file 
	sheet['B' + str(Kcore_Value + 3)] = time.time()
	G = nx.read_edgelist(FILE_PATH)
	sheet['C' + str(Kcore_Value + 3)] = time.time()
	
	#perform kcore search
	H = nx.k_core (G, int(sys.argv[2]))
	if (not H.nodes()):
		sheet['J' + str(Kcore_Value + 3)] = "The community with K of value: " + sys.argv[2] + " is empty; execution stopped"
		sys.exit(0)
	#record time
	sheet['D' + str(Kcore_Value + 3)] = time.time()
	
	#perform partition on Kcore
	partition = community.best_partition(H)
	#record time
	sheet['E' + str(Kcore_Value + 3)] = time.time()
	
	
	sorted_recover_nodes = sort_by_neighbor(H, G)
	new_partition = vote_for_node(partition, sorted_recover_nodes, G)
	sheet['F' + str(Kcore_Value + 3)] = time.time()

	#write current time to log, for the calculation of total time
	sheet['H' + str(Kcore_Value + 3)] = time.time()
	print new_partition 

	#calculate the total time spend and write it to the log file
	sheet['I' + str(Kcore_Value + 3)] = sheet['H' + str(Kcore_Value + 3)].value - sheet['B' + str(Kcore_Value + 3)].value
	
	#write the number of nodes in kcore into log file
	sheet['J' + str(Kcore_Value + 3)] = H.number_of_nodes()

	#write modularity of the partitions into log file
	mo = community.modularity(new_partition,G)
	sheet['K' + str(Kcore_Value + 3)] = mo

	#write number of nodes in each ration range into log file
	sheet['L' + str(Kcore_Value + 3)] = ratio[4]
	sheet['M' + str(Kcore_Value + 3)] = ratio[3]
	sheet['N' + str(Kcore_Value + 3)] = ratio[2]
	sheet['O' + str(Kcore_Value + 3)] = ratio[1]
	sheet['P' + str(Kcore_Value + 3)] = ratio[0]
	wb.save(FILE_LOG_NAME)
	

